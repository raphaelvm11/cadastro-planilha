#importando bibliotecas que vou usar
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#setando a aparencia padrao do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

#criando classe para o app
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.all_system()

    def layout_config(self):
        self.title("sistema de Cadastro de Produtos")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Ligth", "Dark", "System"], command=self.change_apm).place(x=50, y=460)


    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de Produtos", font=("Century Gothic Bold", 24), text_color="#fff").place(x=190, y=10)
        spam = ctk.CTkLabel(self, text="Preencha todos os campos do formulário", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha=ficheiro.active
            folha['A1']="Nome"
            folha['B1']="Contato"
            folha['C1']="Idade"
            folha['D1']="Genero"
            folha['E1']="Email"
            folha['F1']="Observação"

            ficheiro.save("clientes.xlsx")


        def submit():

            #pegando os dados dos entrys
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            adress = adress_value.get()
            obs = obs_entry.get(0.0, END)

            if (name =="" or contact == "" or age =="" or adress ==""):
                messagebox.showerror("Sistema", "ERRO!\nPor favor preencha todos os dados")
            else:

                ficheiro = openpyxl.load_workbook("clientes.xlsx")
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=adress)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save("clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso")


        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            adress_value.set("")
            obs_entry.delete(0.0, END)

        #text variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        adress_value = StringVar()



        #entrys(espaços para digitar)
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        adress_entry = ctk.CTkEntry(self, width=200, textvariable=adress_value, font=("Century Gothic Bold", 16), fg_color="transparent")


        #combobox(definir genero)
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic Bold", 14))
        gender_combobox.set("Masculino")

        #entrada de observação
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        #labels(nomes para denominar as entrys)
        lb_name = ctk.CTkLabel(self, text="Nome", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])
        lb_adress = ctk.CTkLabel(self, text="Email", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic Bold", 24), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_submit = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420) 

        #posicionando na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)
        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)
        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)
        lb_adress.place(x=50, y=190)
        adress_entry.place(x=50, y=220)
        lb_obs.place(x=25, y=260)
        obs_entry.place(x=170, y=260)



    def change_apm(self, new_appearence_mode):
        ctk.set_appearance_mode(new_appearence_mode)






#abrir janela
if __name__=="__main__":
    app = App()
    app.mainloop()